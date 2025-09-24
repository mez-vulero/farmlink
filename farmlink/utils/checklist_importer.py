# your_app/your_app/doctype_utils/checklist_importer.py
import json
import frappe
from frappe.utils.file_manager import get_file_path
from openpyxl import load_workbook

# Allowed enums in your Template Item
ALLOWED_ITEM_KIND = {"Indicator", "Document", "Procedure", "Other"}
ALLOWED_EVIDENCE_MODE = {"Always", "On Yes", "On No", "On Partial", "Never"}
ALLOWED_SEVERITY = {
    "Critical", "Major", "Minor", "Info", "",
    "European Union Deforestation Regulation Requirements",
    "Core Criterion",
    "Improvement Criterion",
    "Mandatory Smart Meter Base Level"
}

def _norm(s):
    return (s or "").strip()

def _norm_title(s):
    return _norm(s).title()

def _resolve_header_map(ws, header_row):
    """Return dict: header_text -> column_index (1-based)"""
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        txt = str(v).strip()
        if txt:
            headers[txt] = col
    return headers

def _pick(row, headers, key_label):
    """Read cell by header label (exact match)."""
    col = headers.get(key_label)
    if not col:
        return ""
    v = row[col]
    return v.value if v is not None else ""

def _pick_by_idx(row, idx):
    """1-based column index reader."""
    if not idx:
        return ""
    v = row[idx]
    return v.value if v is not None else ""

def _normalize_value(kind, raw, norm_map=None):
    s = _norm(str(raw))
    if not s:
        return ""
    # optional custom normalization mapping from mapping_json
    if norm_map and s.lower() in norm_map:
        return norm_map[s.lower()]
    if kind == "evidence_mode":
        # normalize common variants
        m = {
            "always": "Always",
            "on yes": "On Yes",
            "on no": "On No",
            "on partial": "On Partial",
            "never": "Never",
        }
        return m.get(s.lower(), s)
    if kind == "item_kind":
        m = {
            "indicator": "Indicator",
            "document": "Document",
            "procedure": "Procedure",
            "other": "Other",
        }
        return m.get(s.lower(), s)
    if kind == "severity":
        m = {
            "critical": "Critical",
            "major": "Major",
            "minor": "Minor",
            "info": "Info",
            "": "",
        }
        return m.get(s.lower(), _norm_title(s))
    return s

def _assert_enum(value, allowed, field_label, rownum):
    if value and value not in allowed:
        frappe.throw(f"Row {rownum}: invalid {field_label} '{value}'. Allowed: {', '.join(sorted(allowed))}.")

def _get_or_create_template(scheme_version, checklist_type, title, status="Active"):
    # Unique by (scheme_version, checklist_type, title)
    name = frappe.db.get_value(
        "Checklist Template",
        {"scheme_version": scheme_version, "checklist_type": checklist_type, "title": title},
        "name",
    )
    if name:
        return frappe.get_doc("Checklist Template", name), False
    doc = frappe.get_doc({
        "doctype": "Checklist Template",
        "scheme_version": scheme_version,
        "checklist_type": checklist_type,   # Select string
        "title": title,
        "status": status,
    })
    doc.insert(ignore_permissions=True)
    return doc, True

@frappe.whitelist()
def preview_checklist_from_excel(file_url: str, mapping_json: str, limit: int = 15):
    """
    Preview the mapped rows (no DB writes).
    mapping_json supports:
      {
        "sheet_name": "Sheet1",
        "header_row": 1,
        "columns": { "section": "Section", "subsection": "Sub-Section", "requirement_code":"Code", "requirement_text":"Requirement", "guidance":"Guidance", "severity":"Severity", "evidence_mode":"Evidence Mode", "item_kind":"Item Kind", "conditional_rule":"Conditional" },
        "columns_by_index": { "section":1, "subsection":2, "requirement_code":3, "requirement_text":4, "severity":5 },
        "defaults": { "item_kind":"Indicator", "evidence_mode":"On Yes" },
        "normalize": {
           "evidence_mode": { "always":"Always", "on yes":"On Yes" },
           "severity": { "critical":"Critical" }
        },
        "drop_if_blank": ["requirement_text"]
      }
    """
    mapping = json.loads(mapping_json)
    file_path = get_file_path(file_url)
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[mapping["sheet_name"]]

    header_row = int(mapping.get("header_row", 1))
    headers = _resolve_header_map(ws, header_row)
    use_idx = mapping.get("columns_by_index") or {}
    use_hdr = mapping.get("columns") or {}
    defaults = mapping.get("defaults", {})
    norm = mapping.get("normalize", {})
    drop_if_blank = set(mapping.get("drop_if_blank", []))

    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {c: ws.cell(row=r, column=c) for c in range(1, ws.max_column + 1)}

        def getv(field, kind=None):
            if field in use_idx:
                raw = _pick_by_idx(row, int(use_idx[field]))
            elif field in use_hdr:
                raw = _pick(row, headers, use_hdr[field])
            else:
                raw = ""
            return _normalize_value(kind or field, raw, norm.get(field))

        # map core fields
        item = {
            "section": getv("section"),
            "subsection": getv("subsection"),
            "requirement_code": getv("requirement_code"),
            "requirement_text": getv("requirement_text"),
            "guidance": getv("guidance"),
            "severity": getv("severity", "severity") or _normalize_value("severity", defaults.get("severity", ""), norm.get("severity")),
            "item_kind": getv("item_kind", "item_kind") or defaults.get("item_kind", "Indicator"),
            "evidence_mode": getv("evidence_mode", "evidence_mode") or defaults.get("evidence_mode", "Never"),
            "conditional_rule": getv("conditional_rule"),
        }

        # drop empty lines
        if any(k in drop_if_blank and not _norm(item.get(k)) for k in drop_if_blank):
            continue
        if not (item["requirement_text"] or item["requirement_code"] or item["section"]):
            continue

        # validate enums (preview-time)
        _assert_enum(item["item_kind"], ALLOWED_ITEM_KIND, "Item Kind", r)
        _assert_enum(item["evidence_mode"], ALLOWED_EVIDENCE_MODE, "Evidence Mode", r)
        _assert_enum(item["severity"], ALLOWED_SEVERITY, "Severity", r)

        out.append(item)
        if len(out) >= limit:
            break

    return {"rows": out, "preview_count": len(out)}

@frappe.whitelist()
def import_checklist_from_excel(
    file_url: str,
    scheme_version: str,
    checklist_type: str,         # Select string: Standard / Individual / Farm Group
    title: str,
    mapping_json: str,
    status: str = "Active",
    mode: str = "replace"        # replace | append
):
    """
    Import a Checklist Template + Items from an Excel into your current schema.
    - Respects both header-based and index-based mappings.
    - Validates enums (item_kind, evidence_mode, severity).
    - 'replace' wipes current items; 'append' adds new ones (no dedup by code unless you add it).
    """
    mapping = json.loads(mapping_json)
    file_path = get_file_path(file_url)
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[mapping["sheet_name"]]

    header_row = int(mapping.get("header_row", 1))
    headers = _resolve_header_map(ws, header_row)
    use_idx = mapping.get("columns_by_index") or {}
    use_hdr = mapping.get("columns") or {}
    defaults = mapping.get("defaults", {})
    norm = mapping.get("normalize", {})
    drop_if_blank = set(mapping.get("drop_if_blank", ["requirement_text"]))

    # Get or create the template record
    tmpl, created = _get_or_create_template(scheme_version, checklist_type, title, status=status)
    if not created:
        tmpl = frappe.get_doc("Checklist Template", tmpl.name)
        # status may change
        tmpl.status = status

    # Clear items if replace
    if mode == "replace":
        tmpl.items = []

    sort_order = 1 if mode == "replace" else (len(tmpl.items) + 1)

    for r in range(header_row + 1, ws.max_row + 1):
        row = {c: ws.cell(row=r, column=c) for c in range(1, ws.max_column + 1)}

        def getv(field, kind=None):
            if field in use_idx:
                raw = _pick_by_idx(row, int(use_idx[field]))
            elif field in use_hdr:
                raw = _pick(row, headers, use_hdr[field])
            else:
                raw = ""
            return _normalize_value(kind or field, raw, norm.get(field))

        item = {
            "doctype": "Checklist Template Item",
            "section": getv("section"),
            "subsection": getv("subsection"),
            "requirement_code": getv("requirement_code"),
            "requirement_text": getv("requirement_text"),
            "guidance": getv("guidance"),
            "severity": getv("severity", "severity") or _normalize_value("severity", defaults.get("severity", ""), norm.get("severity")),
            "item_kind": getv("item_kind", "item_kind") or defaults.get("item_kind", "Indicator"),
            "evidence_mode": getv("evidence_mode", "evidence_mode") or defaults.get("evidence_mode", "Never"),
            "conditional_rule": getv("conditional_rule"),
            "sort_order": sort_order,
        }

        # Skip empty lines
        if any(k in drop_if_blank and not _norm(item.get(k)) for k in drop_if_blank):
            continue
        if not (item["requirement_text"] or item["requirement_code"] or item["section"]):
            continue

        # Validate enums
        _assert_enum(item["item_kind"], ALLOWED_ITEM_KIND, "Item Kind", r)
        _assert_enum(item["evidence_mode"], ALLOWED_EVIDENCE_MODE, "Evidence Mode", r)
        _assert_enum(item["severity"], ALLOWED_SEVERITY, "Severity", r)

        tmpl.append("items", item)
        sort_order += 1

    if tmpl.is_new():
        tmpl.insert(ignore_permissions=True)
    else:
        tmpl.save(ignore_permissions=True)
    frappe.db.commit()
    return {"template": tmpl.name, "items": len(tmpl.items), "mode": mode}
